import io
from decimal import Decimal

from django.db.models import Sum
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from apps.finance.models import Project
from apps.finance.services import TripleAxisService


def generate_project_excel(project: Project) -> bytes:
    """
    Genera un reporte Excel del proyecto con 4 hojas.

    Returns:
        bytes del archivo XLSX generado.
    """
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    number_format = '#,##0.00'

    # --- Hoja 1: Resumen ---
    ws = wb.active
    ws.title = "Resumen"

    consumed_hours = TripleAxisService.calculate_consumed_hours(project)
    consumption_pct = TripleAxisService.calculate_consumption_percent(project)
    actual_cost = TripleAxisService.calculate_actual_cost(project)

    latest = project.health_snapshots.order_by("-timestamp").first()
    progress_pct = latest.progress_percent if latest else Decimal("0")
    earned_value = latest.earned_value if latest else Decimal("0")

    summary_data = [
        ["Proyecto", project.name],
        ["Codigo", project.code],
        ["Cliente", project.client_name],
        ["Estado", project.current_health_status],
        [""],
        ["Presupuesto (horas)", float(project.budget_hours)],
        ["Presupuesto (monto)", float(project.client_invoice_amount)],
        ["Margen Objetivo (%)", float(project.target_margin)],
        [""],
        ["Horas Consumidas", float(consumed_hours)],
        ["Consumo (%)", float(consumption_pct)],
        ["Costo Real", float(actual_cost)],
        ["Progreso (%)", float(progress_pct)],
        ["Valor Ganado", float(earned_value)],
    ]

    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 1:
                cell.font = Font(bold=True)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 25

    # --- Hoja 2: Time Entries ---
    ws2 = wb.create_sheet("Time Entries")
    headers_te = ["Fecha", "Fase", "Rol", "Usuario", "Horas", "Costo"]

    for col, header in enumerate(headers_te, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    entries = project.time_entries.order_by("-date").select_related("phase", "billing_role")

    for row_idx, entry in enumerate(entries, 2):
        ws2.cell(row=row_idx, column=1, value=entry.date)
        ws2.cell(row=row_idx, column=2, value=entry.phase.name if entry.phase else "-")
        ws2.cell(row=row_idx, column=3, value=entry.billing_role.role_name if entry.billing_role else "-")
        ws2.cell(row=row_idx, column=4, value=entry.user_name)
        cell_h = ws2.cell(row=row_idx, column=5, value=float(entry.duration_hours))
        cell_h.number_format = number_format
        cell_c = ws2.cell(row=row_idx, column=6, value=float(entry.cost))
        cell_c.number_format = number_format

    for col in range(1, 7):
        ws2.column_dimensions[get_column_letter(col)].width = 18

    # --- Hoja 3: Por Fase ---
    ws3 = wb.create_sheet("Por Fase")
    headers_phase = ["Fase", "Estimado (h)", "Real (h)", "Desviacion (h)", "Desviacion (%)"]

    for col, header in enumerate(headers_phase, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for row_idx, phase in enumerate(project.phases.all(), 2):
        actual = (
            phase.time_entries.aggregate(total=Sum("duration_hours"))["total"]
            or Decimal("0")
        )
        deviation = actual - phase.estimated_hours
        deviation_pct = Decimal("0")
        if phase.estimated_hours > 0:
            deviation_pct = (deviation / phase.estimated_hours * 100).quantize(Decimal("0.01"))

        ws3.cell(row=row_idx, column=1, value=phase.name)
        ws3.cell(row=row_idx, column=2, value=float(phase.estimated_hours)).number_format = number_format
        ws3.cell(row=row_idx, column=3, value=float(actual)).number_format = number_format
        ws3.cell(row=row_idx, column=4, value=float(deviation)).number_format = number_format
        ws3.cell(row=row_idx, column=5, value=float(deviation_pct)).number_format = number_format

    for col in range(1, 6):
        ws3.column_dimensions[get_column_letter(col)].width = 18

    # --- Hoja 4: Historico ---
    ws4 = wb.create_sheet("Historico")
    headers_hist = ["Fecha", "Consumo (%)", "Progreso (%)", "Costo", "Valor Ganado", "Estado", "Score"]

    for col, header in enumerate(headers_hist, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    snapshots = project.health_snapshots.order_by("-timestamp")

    for row_idx, snap in enumerate(snapshots, 2):
        ws4.cell(row=row_idx, column=1, value=snap.timestamp.strftime("%Y-%m-%d %H:%M"))
        ws4.cell(row=row_idx, column=2, value=float(snap.consumption_percent)).number_format = number_format
        ws4.cell(row=row_idx, column=3, value=float(snap.progress_percent)).number_format = number_format
        ws4.cell(row=row_idx, column=4, value=float(snap.budget_consumed)).number_format = number_format
        ws4.cell(row=row_idx, column=5, value=float(snap.earned_value)).number_format = number_format
        ws4.cell(row=row_idx, column=6, value=snap.health_status)
        ws4.cell(row=row_idx, column=7, value=snap.health_score)

    for col in range(1, 8):
        ws4.column_dimensions[get_column_letter(col)].width = 18

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
